import pandas as pd
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font,PatternFill


def add_table(dataframe,workbook,start_row_initial,font_style,fill_style,report_name):

    start_row_final = start_row_initial

    with pd.ExcelWriter(workbook, engine= "openpyxl") as writer:
        dataframe.to_excel(writer,sheet_name= "Sheet 1",startrow=start_row_final,index = False)
        start_row_final += len(dataframe) + 2

    wb = load_workbook(workbook)
    ws = wb.active
        
    # Rewrite the header row with the font style
    header_values = dataframe.columns.tolist()  # Get the column names
    for col_num, value in enumerate(header_values, start=1):  # start=1 for Excel 1-based index
        cell = ws.cell(row=start_row_initial+1, column=col_num)  # Row 2 for headers
        cell.value = value  # Set the header text
        cell.font = font_style  # Apply the font style
        cell.fill = fill_style

    merge_row_start = start_row_initial - 1

    ws.merge_cells(start_row = merge_row_start,end_row= merge_row_start,start_column=1,end_column=len(header_values))
    ws.cell(column = 1,row = merge_row_start).value = report_name
    ws.cell(column = 1,row = merge_row_start).font = font_style
    ws.cell(column = 1,row = merge_row_start).fill = fill_style

    wb.save(workbook)

    return start_row_final,wb


if __name__ == "__main__":

    # Set random seed for reproducibility
    np.random.seed(42)

    # Generate sample data
    num_samples = 100
    names = [f'Person {i+1}' for i in range(num_samples)]
    ages = np.random.randint(18, 65, size=num_samples)
    genders = np.random.choice(['Male', 'Female'], size=num_samples)
    salaries = np.round(np.random.normal(loc=50000, scale=15000, size=num_samples), 2)

    # Create a DataFrame
    data = {
        'Name': names,
        'Age': ages,
        'Gender': genders,
        'Salary': salaries
    }

    df = pd.DataFrame(data)

    font_style = Font(name = 'Arial',size = 12, bold = True, color = 'FF000000')
    fill_style = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')

    output = add_table(df,'test_final.xlsx',2,font_style,fill_style,"Sample Report 1")